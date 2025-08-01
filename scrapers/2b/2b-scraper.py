from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Chrome Setup ===
options = Options()
# options.add_argument("--headless=new")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")

# === Input Excel ===
input_excel = "2b-targets.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "2b-outputs"
os.makedirs(output_dir, exist_ok=True)

# === Final Output File (One Workbook) ===
timestamp = datetime.now().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"2b-all-categories_{timestamp}.xlsx")

# === Excel Styling Function (Per Sheet) ===
def style_sheet(ws, category_name, date_str):
    """Apply consistent styling: for header, centered, URL column 30 width, no wrap."""
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    # Insert and merge header row (Row 1)
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = f"2B {category_name} {date_str}"
    merged_cell.font = header_font
    merged_cell.fill = header_fill
    merged_cell.alignment = center_align

    # Style header row (Row 2)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Style body and set column widths
    for column in ws.columns:
        col_letter = get_column_letter(column[0].column)
        header = column[1].value if len(column) > 1 else None

        if header == "Product URL":
            # Fixed width, left-aligned, no wrap, shrink to fit
            ws.column_dimensions[col_letter].width = 30
            for cell in column[2:]:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True  # Changed to True to match Option 1
                )
                cell.font = body_font
                cell.border = border
        else:
            # Auto width for other columns
            max_length = 0
            for cell in column[2:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 5

            # Center-align non-URL cells
            for cell in column[2:]:
                cell.alignment = center_align
                cell.font = body_font
                cell.border = border

# === Helper Functions ===
def normalize_price(text):
    if not text:
        return None
    return int(re.sub(r"[^\d]", "", text))

def extract_sku(name):
    if not name:
        return ""
    name = name.upper().replace("\u200f", " ")
    pattern = r'(?:-\s*)?([A-Z0-9][A-Z0-9\s\+\-]{2,})$'
    matches = re.findall(pattern, name)
    if not matches:
        pattern2 = r'([A-Z0-9][A-Z0-9\s\+\-]{2,})'
        matches = re.findall(pattern2, name)
    candidates = [m.strip() for m in matches if any(c.isalpha() for c in m) and len(m.strip()) >= 3]
    return candidates[-1] if candidates else ""

def normalize_sku(sku):
    return re.sub(r'[^a-zA-Z0-9]', '', sku).lower() if sku else ""

# === Start Browser ===
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)

# === Collect Data by Category ===
print("🚀 Starting 2B Scraper...")
data_by_category = {}

for category, url in category_links:
    print(f"\n➡️ Scraping Category: {category}")
    print(f"🔗 {url}")
    driver.get(url)
    time.sleep(2)

    # Infinite Scroll
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # Extract Products
    products = driver.find_elements(By.CSS_SELECTOR, "div.product-item-info")
    print(f"✅ Found {len(products)} products.")

    data = []
    for product in products:
        try:
            title_el = product.find_element(By.CSS_SELECTOR, "a.product-item-link")
            title = title_el.text.strip()
            product_url = title_el.get_attribute("href").strip()

            # New Price
            try:
                new_price_el = product.find_element(By.CSS_SELECTOR, ".special-price .price")
                new_price = normalize_price(new_price_el.text)
            except:
                try:
                    new_price_el = product.find_element(By.CSS_SELECTOR, ".price-box .price")
                    new_price = normalize_price(new_price_el.text)
                except:
                    new_price = None

            # Old Price
            try:
                old_price_el = product.find_element(By.CSS_SELECTOR, ".old-price .price")
                old_price = normalize_price(old_price_el.text)
            except:
                old_price = None

            # SKU Extraction
            product_code = extract_sku(title)
            normalized_code = normalize_sku(product_code)

            data.append({
                "Item Name": title,
                "Old Price": old_price,
                "New Price": new_price,
                "Product Code": product_code,
                "Normalized Code": normalized_code,
                "Product URL": product_url
            })

        except Exception as e:
            print(f"⚠️ Skipped product: {e}")
            continue

    # Save data for this category
    if data:
        data_by_category[category] = data
        print(f"📌 Collected {len(data)} products for '{category}'")
    else:
        print(f"⚠️ No products collected for '{category}'")

# === Save All Data to Single Workbook ===
if data_by_category:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for category, data in data_by_category.items():
            # Clean sheet name (remove invalid chars, limit to 31)
            safe_sheet_name = re.sub(r'[\/\\*?\[\]:"]', '_', category)[:31].strip()
            df_out = pd.DataFrame(data)[[
                "Item Name",
                "Old Price",
                "New Price",
                "Product Code",
                "Normalized Code",
                "Product URL"
            ]]
            df_out.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            style_sheet(writer.sheets[safe_sheet_name], category, datetime.now().strftime("%y-%m-%d"))
    print(f"📁 Saved all categories to: {output_file}")
else:
    print("⚠️ No data collected across all categories.")

# === Cleanup ===
driver.quit()
print("🏁 2B scraping completed.")
