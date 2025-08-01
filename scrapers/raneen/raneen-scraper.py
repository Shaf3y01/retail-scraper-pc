from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, StaleElementReferenceException
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Chrome Setup ===
options = Options()
# options.add_argument('--headless=new')  # Uncomment for headless mode
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)

# === Input Excel ===
input_excel = "raneen-targets.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "raneen-outputs"
os.makedirs(output_dir, exist_ok=True)

# === Final Output File (One Workbook) ===
timestamp = datetime.now().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"raneen-all-categories_{timestamp}.xlsx")

# === Excel Styling Function (Per Sheet) ===
def style_sheet(ws, category_name, date_str):
    """Apply consistent styling: centered body, header, fixed URL column width."""
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    # Insert and merge header row (Row 1)
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = f"Raneen {category_name} {date_str}"
    merged_cell.font = header_font
    merged_cell.fill = header_fill
    merged_cell.alignment = center_align

    # Style header row (Row 2)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Style body and set column widths
    for column in ws.columns:
        col_letter = get_column_letter(column[0].column)
        header = column[1].value if len(column) > 1 else None

        if header == "Product URL":
            # Fixed width (approx 20 chars), left-aligned, shrink to fit
            ws.column_dimensions[col_letter].width = 20
            for cell in column[2:]:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True
                )
                cell.font = body_font
                cell.border = border
        else:
            # Approximate auto-fit by measuring max length
            max_length = 0
            for cell in column[2:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 6

            for cell in column[2:]:
                cell.alignment = center_align
                cell.font = body_font
                cell.border = border

# === Helper Functions ===
def normalize_price(text):
    if not text:
        return None
    return int(re.sub(r"[^\d]", "", text))

def extract_sku(name):
    if not name:
        return ""
    name = name.upper().replace("\u200f", " ")
    pattern = r'(?:-\s*)?([A-Z0-9][A-Z0-9\s\+\-]{2,})$'
    matches = re.findall(pattern, name)
    if not matches:
        pattern2 = r'([A-Z0-9][A-Z0-9\s\+\-]{2,})'
        matches = re.findall(pattern2, name)
    candidates = [m.strip() for m in matches if any(c.isalpha() for c in m) and len(m.strip()) >= 3]
    return candidates[-1] if candidates else ""

def normalize_sku(sku):
    return re.sub(r'[\-_/\\\.\(\)\s]', '', sku).lower() if sku else ""

# === Start Scraping ===
print("🚀 Starting Raneen Scraper")
data_by_category = {}

for category, url in category_links:
    print(f"\n➡️ Scraping Category: {category} | URL: {url}")
    driver.get(url)
    time.sleep(2)

    # Load all products using infinite scroll
    prev_count = -1
    same_count_repeats = 0
    max_repeats = 3

    while same_count_repeats < max_repeats:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        products = driver.find_elements(By.CSS_SELECTOR, "div.product-item-info")
        current_count = len(products)
        print(f"🔍 Products loaded so far: {current_count}")
        if current_count == prev_count:
            same_count_repeats += 1
        else:
            same_count_repeats = 0
            prev_count = current_count

    product_cards = driver.find_elements(By.CSS_SELECTOR, "div.product-item-info")
    print(f"✅ Total products loaded: {len(product_cards)}")

    data = []
    for card in product_cards:
        try:
            title_el = card.find_element(By.CSS_SELECTOR, "a.product-item-link")
            title = title_el.text.strip()
            product_url = title_el.get_attribute("href").strip()

            # Unified Price Logic
            try:
                price_box = card.find_element(By.CSS_SELECTOR, ".price-box.price-final_price")
                new_price = old_price = None

                # Case 1: Special + Old
                special_els = price_box.find_elements(By.CSS_SELECTOR, ".special-price .price-wrapper")
                old_els = price_box.find_elements(By.CSS_SELECTOR, ".old-price .price-wrapper")
                if special_els:
                    new_price = normalize_price(special_els[0].text)
                    if old_els:
                        old_price = normalize_price(old_els[0].text)
                else:
                    # Case 2: Regular price wrapper
                    reg_els = price_box.find_elements(By.CSS_SELECTOR, ".price-container .price-wrapper")
                    if reg_els:
                        new_price = normalize_price(reg_els[0].text)
                    else:
                        # Case 3: Fallback to raw spans
                        curr_els = price_box.find_elements(By.CSS_SELECTOR, ".current-price")
                        old_els = price_box.find_elements(By.CSS_SELECTOR, ".old-price")
                        if curr_els:
                            new_price = normalize_price(curr_els[0].text)
                        if old_els:
                            old_price = normalize_price(old_els[0].text)
            except:
                new_price = None
                old_price = None

            # SKU Extraction
            product_code = extract_sku(title)
            normalized_code = normalize_sku(product_code)

            data.append({
                "Item Name": title,
                "Old Price": old_price,
                "New Price": new_price,
                "Product Code": product_code,
                "Normalized Code": normalized_code,
                "Product URL": product_url
            })

        except Exception as e:
            print(f"⚠️ Skipping product: {e}")

    # Save data for this category
    if data:
        data_by_category[category] = data
        print(f"📌 Collected {len(data)} products for '{category}'")
    else:
        print(f"⚠️ No products collected for '{category}'")

# === Save All Data to Single Workbook ===
if data_by_category:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for category, data in data_by_category.items():
            safe_sheet_name = re.sub(r'[^\w\s-]', '_', category)[:31].strip()
            df_out = pd.DataFrame(data)[[
                "Item Name",
                "Old Price",
                "New Price",
                "Product Code",
                "Normalized Code",
                "Product URL"
            ]]
            df_out.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            style_sheet(writer.sheets[safe_sheet_name], category, datetime.now().strftime("%y-%m-%d"))
    print(f"📁 Saved all categories to: {output_file}")
else:
    print("⚠️ No data collected across all categories.")

# === Cleanup ===
driver.quit()
print("🏁 Raneen scraping completed.")
