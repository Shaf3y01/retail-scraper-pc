from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Chrome Setup ===
options = Options()
# options.add_argument('--headless=new')  # Uncomment for headless mode
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_experimental_option('useAutomationExtension', False)
driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd('Network.setUserAgentOverride', {
    "userAgent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
})
wait = WebDriverWait(driver, 10)

# === Input Excel ===
input_excel = "btech-targets.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "btech-outputs"
os.makedirs(output_dir, exist_ok=True)

# === Final Output File (One Workbook) ===
timestamp = datetime.now().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"btech-all-categories_{timestamp}.xlsx")

# === Excel Styling Function (Per Sheet) ===
def style_sheet(ws, category_name, date_str):
    """Apply consistent styling: black header, centered, URL column width 20, shrink-to-fit."""
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    # Insert and merge header row (Row 1)
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = f"Btech {category_name} {date_str}"
    merged_cell.font = header_font
    merged_cell.fill = header_fill
    merged_cell.alignment = center_align

    # Style header row (Row 2)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Style body and set column widths
    for column in ws.columns:
        col_letter = get_column_letter(column[0].column)
        header = column[1].value if len(column) > 1 else None

        if header == "Product URL":
            # Fixed width, left-aligned, shrink to fit
            ws.column_dimensions[col_letter].width = 20
            for cell in column[2:]:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True
                )
                cell.font = body_font
                cell.border = border
        else:
            # Auto width for other columns
            max_length = 0
            for cell in column[2:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 6

            # Center-align non-URL cells
            for cell in column[2:]:
                cell.alignment = center_align
                cell.font = body_font
                cell.border = border

# === Helper Functions ===
def normalize_price(text):
    """Extract integer from price text."""
    if not text:
        return None
    return int(re.sub(r"[^\d]", "", text))

def extract_sku(name):
    """Btech-style SKU extraction."""
    if not name:
        return ""
    name = name.upper().replace("\u200f", " ")
    pattern = r'(?:-\s*)?([A-Z0-9][A-Z0-9\s\+\-]{2,})$'
    matches = re.findall(pattern, name)
    if not matches:
        pattern2 = r'([A-Z0-9][A-Z0-9\s\+\-]{2,})'
        matches = re.findall(pattern2, name)
    candidates = [m.strip() for m in matches if any(c.isalpha() for c in m) and len(m.strip()) >= 3]
    return candidates[-1] if candidates else ""

def normalize_sku(sku):
    """Remove separators and lowercase."""
    return re.sub(r'[^a-zA-Z0-9]', '', sku).lower() if sku else ""

# === Extract Total Expected Products ===
def extract_total_expected_products(driver):
    """Extract total from #product-search-item-count."""
    try:
        el = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span#product-search-item-count"))
        )
        text = el.text.strip()
        print(f"🔍 Found product count: {text}")
        if text.isdigit():
            return int(text)
        numbers = re.findall(r'\d+', text)
        return int(numbers[0]) if numbers else None
    except Exception as e:
        print(f"❌ Could not find product count: {str(e)}")
        return None

# === Start Scraping ===
print("🚀 Starting Btech Scraper")
data_by_category = {}

for category, url in category_links:
    print(f"\n➡️ Scraping Category: {category}")
    print(f"🔗 URL: {url}")
    driver.get(url)
    time.sleep(3)

    # Wait for product container
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.products.wrapper.grid.products-grid")))
    except TimeoutException:
        print("❌ Timeout: Product container not found. Skipping.")
        continue

    # Get expected total
    expected_total = extract_total_expected_products(driver)
    if not expected_total:
        print("⚠️ Could not determine product count. Skipping category.")
        continue
    max_scrape_limit = expected_total + 2
    print(f"📊 Expected: {expected_total} | Max: {max_scrape_limit}")

    # Click "Load More" safely
    previous_count = 0
    click_count = 0
    max_clicks = (expected_total // 30) + 5  # 30 per click + margin

    while click_count < max_clicks:
        time.sleep(2)
        current_cards = driver.find_elements(By.CSS_SELECTOR, "div.plpContentWrapper")
        current_count = len(current_cards)
        print(f"🔄 Loaded {current_count} products...")

        if current_count >= max_scrape_limit:
            print(f"✅ Reached limit: {current_count}")
            break

        if current_count == previous_count:
            click_count += 1
        else:
            click_count = 0

        previous_count = current_count

        try:
            load_more = driver.find_element(By.CSS_SELECTOR, "div.amscroll-load-button")
            if "عرض" in load_more.text or "المزيد" in load_more.text:
                print(f"➡️ Clicking 'Load More' ({click_count + 1}/{max_clicks})...")
                try:
                    load_more.click()
                except:
                    driver.execute_script("arguments[0].click();", load_more)
                time.sleep(3)
        except:
            print("🔚 'Load More' not found.")
            break

    # Final product list
    product_cards = driver.find_elements(By.CSS_SELECTOR, "div.plpContentWrapper")
    print(f"✅ Final count: {len(product_cards)}")

    # === Parse Products ===
    data = []
    wrappers = driver.find_elements(By.CSS_SELECTOR, "a.listingWrapperSection")
    print(f"Found {len(wrappers)} products to parse")

    for wrapper in wrappers:
        try:
            # Title
            title_els = wrapper.find_elements(By.CSS_SELECTOR, "h2.plpTitle")
            if not title_els or not title_els[0].text.strip():
                continue
            title = title_els[0].text.strip()

            # URL (from wrapper anchor)
            product_url = wrapper.get_attribute("href").strip()

            # New Price
            try:
                new_price_el = wrapper.find_element(By.CSS_SELECTOR, "span.special-price span.price-wrapper")
                new_price = normalize_price(new_price_el.text)
            except:
                new_price = None

            # Old Price
            try:
                old_price_el = wrapper.find_element(By.CSS_SELECTOR, "span.old-price.was-price span.price-wrapper")
                old_price = normalize_price(old_price_el.text)
            except:
                old_price = None

            # SKU Extraction
            product_code = extract_sku(title)
            normalized_code = normalize_sku(product_code)

            data.append({
                "Item Name": title,
                "Old Price": old_price,
                "New Price": new_price,
                "Product Code": product_code,
                "Normalized Code": normalized_code,
                "Product URL": product_url
            })
        except Exception as e:
            print(f"❌ Skipped product: {e}")
            continue

    # Save data for this category
    if data:
        data_by_category[category] = data
        print(f"📌 Collected {len(data)} products for '{category}'")
    else:
        print(f"⚠️ No products collected for '{category}'")

# === Save All Data to Single Workbook ===
if data_by_category:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for category, data in data_by_category.items():
            safe_sheet_name = re.sub(r'[^\w\s-]', '_', category)[:31].strip()
            df_out = pd.DataFrame(data)[[
                "Item Name",
                "Old Price",
                "New Price",
                "Product Code",
                "Normalized Code",
                "Product URL"
            ]]
            df_out.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            style_sheet(writer.sheets[safe_sheet_name], category, datetime.now().strftime("%y-%m-%d"))
    print(f"📁 Saved all categories to: {output_file}")
else:
    print("⚠️ No data collected across all categories.")

# === Cleanup ===
driver.quit()
print("🏁 Btech scraping completed.")
