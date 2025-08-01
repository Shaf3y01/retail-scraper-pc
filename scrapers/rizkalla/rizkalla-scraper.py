from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Chrome Setup ===
options = Options()
# options.add_argument('--headless=new')  # Uncomment for headless mode
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_experimental_option('useAutomationExtension', False)
driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd('Network.setUserAgentOverride', {
    "userAgent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
})
wait = WebDriverWait(driver, 10)

# === Input Excel ===
input_excel = "rizkalla-targets.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "rizkalla-outputs"
os.makedirs(output_dir, exist_ok=True)

# === Final Output File (One Workbook) ===
timestamp = datetime.now().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"rizkalla-all-categories_{timestamp}.xlsx")

# === Excel Styling Function (Per Sheet) ===
def style_sheet(ws, category_name, date_str):
    """Apply consistent styling: black header, centered, URL column 30 width, no wrap."""
    header_fill = PatternFill(start_color="191970", end_color="191970", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    # Insert and merge header row (Row 1)
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = f"Rizkalla {category_name} {date_str}"
    merged_cell.font = header_font
    merged_cell.fill = header_fill
    merged_cell.alignment = center_align

    # Style header row (Row 2)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Style body and set column widths
    for column in ws.columns:
        col_letter = get_column_letter(column[0].column)
        header = column[1].value if len(column) > 1 else None

        if header == "Product URL":
            # Fixed width, left-aligned, no wrap, shrink to fit
            ws.column_dimensions[col_letter].width = 30
            for cell in column[2:]:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True  # Changed to True to match Option 1
                )
                cell.font = body_font
                cell.border = border
        else:
            # Auto width for other columns
            max_length = 0
            for cell in column[2:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 6

            # Center-align non-URL cells
            for cell in column[2:]:
                cell.alignment = center_align
                cell.font = body_font
                cell.border = border
                
# === Helper Functions ===
def normalize_price(text):
    """Extracts integer from price text."""
    if not text:
        return None
    match = re.search(r'[\d,]+(?=\.\d+|$)', text)
    if not match:
        return None
    cleaned = match.group(0).replace(',', '')
    return int(cleaned) if cleaned.isdigit() else None

def extract_sku(name):
    """Btech-style SKU extraction."""
    if not name:
        return ""
    name = name.upper().replace("\u200f", " ")
    pattern = r'(?:-\s*)?([A-Z0-9][A-Z0-9\s\+\-]{2,})$'
    matches = re.findall(pattern, name)
    if not matches:
        pattern2 = r'([A-Z0-9][A-Z0-9\s\+\-]{2,})'
        matches = re.findall(pattern2, name)
    candidates = [m.strip() for m in matches if any(c.isalpha() for c in m) and len(m.strip()) >= 3]
    return candidates[-1] if candidates else ""

def normalize_sku(sku):
    """Removes separators and converts to lowercase."""
    return re.sub(r'[^a-zA-Z0-9]', '', sku).lower() if sku else ""

# === Detect Page Type ===
def is_search_page(driver):
    return 'search' in driver.current_url.lower() or 'q=' in driver.current_url.lower()

# === Extract Total Product Count (Fallback Logic) ===
def get_total_product_count(driver):
    counts = []
    # Try 1: div.products-showing (category)
    try:
        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.products-showing#products-showing")))
        text = elem.text.strip()
        numbers = [int(num) for num in re.findall(r'\d+', text)]
        counts.extend(numbers)
    except:
        pass
    # Try 2: h3.search-results_title (search)
    try:
        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "h3.search-results_title")))
        text = elem.text.strip()
        numbers = [int(num) for num in re.findall(r'\d+', text)]
        counts.extend(numbers)
    except:
        pass
    return max(counts) if counts else None

# === Get Product Cards Based on Page Type ===
def get_product_cards(driver):
    if is_search_page(driver):
        return driver.find_elements(By.CSS_SELECTOR, ".search-results_inner > .product-product-grid > product-card")
    else:
        return driver.find_elements(By.CSS_SELECTOR, "div#main-collection-product-grid > product-card")

# === Start Scraping ===
print("🚀 Starting Rizkalla Scraper")
data_by_category = {}

for category, url in category_links:
    print(f"\n➡️ Scraping Category: {category}")
    print(f"🔗 URL: {url}")
    driver.get(url)
    time.sleep(3)

    search_mode = is_search_page(driver)
    products_per_page = 16 if search_mode else 20
    print(f"🔍 Detected mode: {'Search' if search_mode else 'Category'} ({products_per_page} products/page)")

    total_count = get_total_product_count(driver)
    if not total_count:
        print("❌ Could not determine total product count. Skipping category.")
        continue
    print(f"📊 Total products: {total_count}")
    total_pages = (total_count // products_per_page) + (1 if total_count % products_per_page > 0 else 0)
    print(f" totalPages: {total_pages} ({products_per_page} per page)")

    all_data = []
    for page in range(1, total_pages + 1):
        print(f"\n📄 Scraping Page {page} of {total_pages}...")

        try:
            if search_mode:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".search-results_inner")))
            else:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div#main-collection-product-grid")))
        except TimeoutException:
            print("❌ Timeout: Product grid not found.")
            break

        product_cards = get_product_cards(driver)
        print(f"✅ Found {len(product_cards)} product cards on page {page}.")

        page_data = []
        for card in product_cards:
            try:
                title_el = card.find_element(By.CSS_SELECTOR, "section > header > div.product-card_vendor-title > h3 > a")
                title = title_el.text.strip()
                product_url = title_el.get_attribute("href").strip()

                price_container = card.find_element(By.CSS_SELECTOR, "footer div.product-price")

                try:
                    new_price_el = price_container.find_element(By.CSS_SELECTOR, "div.price-sale")
                    new_price = normalize_price(new_price_el.text)
                except:
                    new_price = None

                try:
                    old_price_el = price_container.find_element(By.CSS_SELECTOR, "del.price-compare")
                    old_price = normalize_price(old_price_el.text)
                except:
                    old_price = None

                product_code = extract_sku(title)
                normalized_code = normalize_sku(product_code)

                page_data.append({
                    "Item Name": title,
                    "Old Price": old_price,
                    "New Price": new_price,
                    "Product Code": product_code,
                    "Normalized Code": normalized_code,
                    "Product URL": product_url
                })
            except Exception as e:
                print(f"⚠️ Skipped product: {e}")
                continue

        all_data.extend(page_data)

        # === Pagination: Click "Next" or numbered link ===
        if page < total_pages:
            next_page_num = page + 1
            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, "div.pagination-holder ul.pagination li a.next")
                if next_btn.is_displayed():
                    print("➡️ Clicking 'Next' button...")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
                    time.sleep(1)
                    try:
                        next_btn.click()
                    except:
                        driver.execute_script("arguments[0].click();", next_btn)
                    time.sleep(3)
                    continue
            except:
                pass

            try:
                page_link = driver.find_element(
                    By.CSS_SELECTOR,
                    f"div.pagination-holder ul.pagination li > a[href*='page={next_page_num}']"
                )
                if page_link.is_displayed():
                    print(f"➡️ Clicking page {next_page_num} link...")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", page_link)
                    time.sleep(1)
                    try:
                        page_link.click()
                    except:
                        driver.execute_script("arguments[0].click();", page_link)
                    time.sleep(3)
                    continue
            except:
                print(f"⚠️ Could not navigate to page {next_page_num}")
                break

        print("🔚 Last page reached.")
        break

    # Save data for this category
    if all_data:
        data_by_category[category] = all_data
        print(f"📌 Collected {len(all_data)} products for '{category}'")
    else:
        print(f"⚠️ No products collected for '{category}'")

# === Save All Data to Single Workbook ===
if data_by_category:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for category, data in data_by_category.items():
            safe_sheet_name = re.sub(r'[^\w\s-]', '_', category)[:31].strip()
            df_out = pd.DataFrame(data)[[
                "Item Name",
                "Old Price",
                "New Price",
                "Product Code",
                "Normalized Code",
                "Product URL"
            ]]
            df_out.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            style_sheet(writer.sheets[safe_sheet_name], category, datetime.now().strftime("%y-%m-%d"))
    print(f"📁 Saved all categories to: {output_file}")
else:
    print("⚠️ No data collected across all categories.")

# === Cleanup ===
driver.quit()
print("🏁 Rizkalla scraping completed.")
