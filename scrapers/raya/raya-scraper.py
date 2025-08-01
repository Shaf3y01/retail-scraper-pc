from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Chrome Setup ===
options = Options()
# options.add_argument('--headless=new')  # Uncomment for headless mode
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_experimental_option('useAutomationExtension', False)
driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd('Network.setUserAgentOverride', {
    "userAgent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
})
wait = WebDriverWait(driver, 10)

# === Input Excel ===
input_excel = "raya-targets.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "raya-outputs"
os.makedirs(output_dir, exist_ok=True)

# === Final Output File (One Workbook, Multiple Sheets) ===
timestamp = datetime.now().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"raya-all-categories_{timestamp}.xlsx")

# === Excel Styling Function (Per Sheet) ===
def style_sheet(ws, category_name, date_str):
    """Apply consistent styling: blue header, centered, URL column 30 width, no wrap."""
    header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    # Insert and merge header row (Row 1)
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = f"RAYA {category_name} {date_str}"
    merged_cell.font = header_font
    merged_cell.fill = header_fill
    merged_cell.alignment = center_align

    # Style header row (Row 2)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Style body and set column widths
    for column in ws.columns:
        col_letter = get_column_letter(column[0].column)
        header = column[1].value if len(column) > 1 else None

        if header == "Product URL":
            # Fixed width, left-aligned, no wrap, shrink to fit
            ws.column_dimensions[col_letter].width = 30
            for cell in column[2:]:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True  # Changed to True to match Option 1
                )
                cell.font = body_font
                cell.border = border
        else:
            # Auto width for other columns
            max_length = 0
            for cell in column[2:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 6

            # Center-align non-URL cells
            for cell in column[2:]:
                cell.alignment = center_align
                cell.font = body_font
                cell.border = border
                
# === Helper Functions ===
def normalize_price(text):
    """Extracts integer from price text (removes commas)."""
    if not text:
        return None
    cleaned = re.sub(r'[^\d]', '', text)
    return int(cleaned) if cleaned else None

def extract_sku(name):
    """Btech-style SKU extraction."""
    if not name:
        return ""
    name = name.upper().replace("\u200f", " ")
    pattern = r'(?:-\s*)?([A-Z0-9][A-Z0-9\s\+\-]{2,})$'
    matches = re.findall(pattern, name)
    if not matches:
        pattern2 = r'([A-Z0-9][A-Z0-9\s\+\-]{2,})'
        matches = re.findall(pattern2, name)
    candidates = [m.strip() for m in matches if any(c.isalpha() for c in m) and len(m.strip()) >= 3]
    return candidates[-1] if candidates else ""

def normalize_sku(sku):
    """Removes separators and converts to lowercase."""
    return re.sub(r'[^a-zA-Z0-9]', '', sku).lower() if sku else ""

# === Get Total Product Count (Dual Verification) ===
def get_total_product_count(driver):
    """
    Extracts total count from two sources:
    1. <h3>82 Product</h3>
    2. <p>Showing 82 out of 82</p>
    Returns the largest number found.
    """
    counts = []
    # Source 1: h3 with product count
    try:
        h3_elem = driver.find_element(By.CSS_SELECTOR, "h3.text-secondary-500.text-sm")
        text = h3_elem.text.strip()
        numbers = [int(n) for n in re.findall(r'\d+', text)]
        counts.extend(numbers)
    except Exception as e:
        print("⚠️ Could not extract count from <h3>:", str(e))
        pass

    # Source 2: p with "Showing X out of Y"
    try:
        p_elem = driver.find_element(By.CSS_SELECTOR, "p.text-secondary-400.text-sm.text-center")
        text = p_elem.text.strip()
        numbers = [int(n) for n in re.findall(r'\d+', text)]
        counts.extend(numbers)
    except Exception as e:
        print("⚠️ Could not extract count from <p>:", str(e))
        pass

    return max(counts) if counts else None

# === Start Scraping ===
print("🚀 Starting Raya Scraper")
data_by_category = {}

for category, url in category_links:
    print(f"\n➡️ Scraping Category: {category}")
    print(f"🔗 URL: {url}")
    driver.get(url)
    time.sleep(3)

    # Wait for product grid
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductsGrid")))
        print("✅ Product grid loaded")
    except TimeoutException:
        print("❌ Timeout: Product grid not found. Skipping category.")
        continue

    # Get expected total count (dual verification)
    total_count = get_total_product_count(driver)
    if not total_count:
        print("❌ Could not determine total product count. Skipping category.")
        continue
    print(f"📊 Expected products: {total_count}")

    # Infinite Scroll with Product Count Verification
    seen_count = 0
    max_wait_attempts = 10
    attempt = 0

    while attempt < max_wait_attempts:
        # Scroll to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)  # Allow time for AJAX load

        # Find current number of product cards
        current_cards = driver.find_elements(By.CSS_SELECTOR, "article.ProductCard")
        current_count = len(current_cards)

        print(f"🔄 Loaded {current_count} / {total_count} products...")

        # ✅ If we've reached or exceeded expected count, stop
        if current_count >= total_count:
            print(f"✅ Loaded all {current_count} expected products.")
            break

        # ⚠️ If no new products loaded in this iteration
        if current_count == seen_count:
            attempt += 1
            print(f"⚠️ No new products. Stagnant attempt {attempt}/{max_wait_attempts}")
        else:
            attempt = 0  # Reset if new products loaded

        seen_count = current_count

        # 🛑 Safety break
        if attempt >= max_wait_attempts:
            print("🛑 Max wait attempts reached. May have missed some products.")
            break
    
    # Final verification
    final_cards = driver.find_elements(By.CSS_SELECTOR, "article.ProductCard")
    print(f"✅ Final product count: {len(final_cards)}")
    if len(final_cards) < total_count:
        print(f"⚠️ Warning: Only {len(final_cards)} out of {total_count} products loaded.")
        
    # Extract product cards
    product_cards = driver.find_elements(By.CSS_SELECTOR, "article.ProductCard")
    print(f"✅ Found {len(product_cards)} product cards.")

    data = []
    for card in product_cards:
        try:
            # Title & URL
            try:
                title_link = card.find_element(By.CSS_SELECTOR, "a.flex.flex-col[href]")
                title_el = title_link.find_element(By.CSS_SELECTOR, "p.name.clamp-text")
                title = title_el.text.strip()
                product_url = "https://www.rayashop.com" + title_link.get_attribute("href").strip()
            except Exception as e:
                print("⚠️ Skipped product: missing title or URL")
                continue

            # New Price
            try:
                new_price_el = card.find_element(By.CSS_SELECTOR, "span.text-primary-500:not(.line-through)")
                new_price_text = new_price_el.text.strip()
                new_price = normalize_price(new_price_text)
            except:
                new_price = None

            # Old Price
            try:
                old_price_el = card.find_element(By.CSS_SELECTOR, "span.line-through")
                old_price_text = old_price_el.text.strip()
                old_price = normalize_price(old_price_text)
            except:
                old_price = None

            # SKU Extraction
            product_code = extract_sku(title)
            normalized_code = normalize_sku(product_code)

            data.append({
                "Item Name": title,
                "Old Price": old_price,
                "New Price": new_price,
                "Product Code": product_code,
                "Normalized Code": normalized_code,
                "Product URL": product_url
            })

        except Exception as e:
            print(f"⚠️ Skipped product: {e}")
            continue

    # Save data for this category
    if data:
        data_by_category[category] = data
        print(f"📌 Collected {len(data)} products for '{category}'")
    else:
        print(f"⚠️ No products collected for '{category}'")

# === Save All Data to Single Workbook ===
if data_by_category:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for category, data in data_by_category.items():
            safe_sheet_name = re.sub(r'[\/\\*?\[\]:]', '_', category)[:31]
            df_out = pd.DataFrame(data)
            df_out = df_out[[
                "Item Name",
                "Old Price",
                "New Price",
                "Product Code",
                "Normalized Code",
                "Product URL"
            ]]
            df_out.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            style_sheet(writer.sheets[safe_sheet_name], category, datetime.now().strftime("%y-%m-%d"))
    print(f"📁 Saved all categories to: {output_file}")
else:
    print("⚠️ No data collected across all categories.")

# === Cleanup ===
driver.quit()
print("🏁 Raya scraping completed.")
