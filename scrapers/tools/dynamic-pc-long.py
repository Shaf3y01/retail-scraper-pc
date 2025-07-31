import os
import re
import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))
RETAILER_FOLDERS = {
    "2b": os.path.join(BASE_DIR, "2b", "2b-outputs"),
    "Btech": os.path.join(BASE_DIR, "btech", "btech-outputs"),
    "Raneen": os.path.join(BASE_DIR, "raneen", "raneen-outputs"),
}
OUTPUT_FOLDER = os.path.join(BASE_DIR, "results", "long")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
REQUIRED_COLUMNS = ["Item Name", "New Price", "Normalized Code", "Product URL"]
HIGHLIGHT_CONFIDENCE_WEAK = 30
HIGHLIGHT_CONFIDENCE_UNMATCHED = 10

def log(msg):
    print(f"[LOG] {msg}")

def extract_info_from_filename(filename):
    match = re.match(r"([a-zA-Z0-9]+)_([A-Za-z0-9\-]+)_\d{4}-\d{2}-\d{2}\.xlsx", filename)
    if match:
        return match.group(1).capitalize(), match.group(2)
    return None, None

def compute_confidence(name1, name2):
    if not name1 or not name2 or pd.isna(name1) or pd.isna(name2):
        return 0.0
    return round(fuzz.token_sort_ratio(str(name1), str(name2)), 2)

def prepare(df, retailer):
    df.columns = df.columns.str.strip()
    if not all(col in df.columns for col in REQUIRED_COLUMNS):
        return None
    df = df[REQUIRED_COLUMNS].copy()
    df["Retailer"] = retailer
    df = df.dropna(subset=["Normalized Code"])
    df["Normalized Code"] = df["Normalized Code"].astype(str).str.strip().str.lower()
    df["Item Name"] = df["Item Name"].astype(str).str.strip()
    df["New Price"] = pd.to_numeric(df["New Price"], errors="coerce")
    return df

def export_results(rows, filename, header_text=None, highlight_confidence=None):
    if not rows:
        log(f"⚠️ No data to export for {filename}")
        return
    valid_rows = [row for row in rows if isinstance(row, pd.DataFrame)]
    if not valid_rows:
        log(f"⚠️ No valid DataFrame rows to export for {filename}")
        return
    try:
        df_out = pd.concat(valid_rows).reset_index(drop=True)
    except Exception as e:
        log(f"❌ Error concatenating rows for {filename}: {e}")
        return
    final_cols = [
        "2B Item Name", "2B Price", "2B Normalized Code",
        "Btech Item Name", "Btech Price", "Btech Normalized Code",
        "Raneen Item Name", "Raneen Price", "Raneen Normalized Code",
        "Confidence", "Best Price", "Lowest Retailer", "Product URL"
    ]
    # Ensure all columns exist before selecting
    for col in final_cols:
        if col not in df_out.columns:
            df_out[col] = None
    # Dynamically fill 'Product URL' with the URL from the retailer with the lowest price
    for idx, row in df_out.iterrows():
        lowest_retailer = row["Lowest Retailer"]
        url_col = f"{lowest_retailer} Product URL"
        url = None
        if url_col in df_out.columns:
            url = row.get(url_col, None)
        df_out.at[idx, "Product URL"] = url
    df_out = df_out[final_cols]
    output_path = os.path.join(OUTPUT_FOLDER, filename)
    # Insert merged header row and column header row
    from openpyxl.utils import get_column_letter
    df_out.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(final_cols))
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = header_text if header_text else filename
    merged_cell.font = Font(bold=True, color="FFFFFF")
    merged_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Style header row (row 2) and body
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Font(bold=False)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_align
            if cell.row == 2:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = body_font
    for column in ws.columns:
        col_idx = column[0].column
        col_letter = get_column_letter(col_idx)
        header = str(column[1].value).strip() if len(column) > 1 and column[1].value else ""
        if "Product URL" in header:
            ws.column_dimensions[col_letter].width = 30
            for cell in column:
                if cell.row > 1:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=False,
                        shrink_to_fit=True
                    )
        else:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column[1:])
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(output_path)
    log(f"✅ Saved to {output_path}")

# === Step 1: Scan retailer folders ===
log("🔍 Scanning retailer folders...")
category_map = {}
for retailer, folder_path in RETAILER_FOLDERS.items():
    log(f"🔎 Checking: {folder_path}")
    if not os.path.exists(folder_path):
        log(f"❌ Folder not found: {folder_path}")
        continue
    files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
    for filename in files:
        ret_name, category = extract_info_from_filename(filename)
        if ret_name and category:
            category_map.setdefault(category, {})[retailer] = os.path.join(folder_path, filename)
log(f"📦 Found {len(category_map)} categories.")
for cat, data in category_map.items():
    log(f"  - {cat}: {list(data.keys())}")

# === Step 2: Process each category ===
matched_categories = []
for category, sources in category_map.items():
    if len(sources) < 2:
        continue
    log(f"\n🚀 Processing category: {category}")
    all_dfs = []
    for retailer, path in sources.items():
        log(f"📥 Reading: {path}")
        try:
            # Read with header at row 2 (index 1) to skip merged header row
            df = pd.read_excel(path, header=1)
            prepared = prepare(df, retailer)
            if prepared is None:
                log(f"⚠️ Skipped {retailer} in {category} — missing required columns")
                continue
            all_dfs.append(prepared)
        except Exception as e:
            log(f"❌ Failed to read {retailer}: {e}")
    if len(all_dfs) < 2:
        log(f"⚠️ Not enough valid data sources for {category}")
        continue
    combined = pd.concat(all_dfs, ignore_index=True)
    grouped = combined.groupby("Normalized Code")
    matched_rows, weak_matched_rows, unmatched_rows = [], [], []
    # Exact matches by Normalized Code
    for code, group in grouped:
        if group["Retailer"].nunique() < 2:
            unmatched_rows.extend(group.to_dict('records'))
            continue
        merged = pd.DataFrame({"Normalized Code": [code]})
        sku_confidences = {}
        for retailer in RETAILER_FOLDERS.keys():
            match = group[group["Retailer"] == retailer]
            if not match.empty:
                merged[f"{retailer} Item Name"] = match.iloc[0]["Item Name"]
                merged[f"{retailer} Old Price"] = match.iloc[0].get("Old Price", None)
                merged[f"{retailer} Price"] = match.iloc[0]["New Price"]
                merged[f"{retailer} Item SKU"] = match.iloc[0].get("Product Code", code)
                merged[f"{retailer} Normalized Code"] = code
                merged[f"{retailer} Product URL"] = match.iloc[0]["Product URL"]
                sku_confidences[retailer] = compute_confidence(match.iloc[0].get("Product Code", code), code)
            else:
                merged[f"{retailer} Item Name"] = "N/A"
                merged[f"{retailer} Old Price"] = None
                merged[f"{retailer} Price"] = pd.NA
                merged[f"{retailer} Item SKU"] = code
                merged[f"{retailer} Normalized Code"] = code
                merged[f"{retailer} Product URL"] = "N/A"
                sku_confidences[retailer] = 0.0
        avg_conf = sum(sku_confidences.values()) / len(sku_confidences)
        merged["Confidence"] = avg_conf
        prices = {r: merged[f"{r} Price"].values[0] for r in RETAILER_FOLDERS if pd.notna(merged[f"{r} Price"].values[0])}
        best_price, best_retailer = (min(prices.values()), min(prices, key=prices.get)) if prices else (None, None)
        merged["Best Price"] = best_price
        merged["Lowest Retailer"] = best_retailer
        if avg_conf >= 81:
            matched_rows.append(merged)
        elif 20 <= avg_conf < 81:
            # If avg_conf < 50, run fuzzy match on Item Name
            if avg_conf < 50:
                pass
            weak_matched_rows.append(merged)
        else:
            # avg_conf < 20, run fuzzy match on Item Name
            # ...fuzzy match logic on Item Name...
            unmatched_rows.append(merged)
    # Fuzzy matching for unmatched items
    new_unmatched_rows = []
    if unmatched_rows:
        unmatched_df = pd.DataFrame(unmatched_rows)
        for retailer in unmatched_df["Retailer"].unique():
            retailer_df = unmatched_df[unmatched_df["Retailer"] == retailer]
            other_df = unmatched_df[unmatched_df["Retailer"] != retailer].reset_index(drop=True)
            for _, row in retailer_df.iterrows():
                if other_df.empty:
                    merged = pd.DataFrame({"Normalized Code": [row["Normalized Code"]]})
                    for r in RETAILER_FOLDERS:
                        if r == retailer:
                            merged[f"{r} Item Name"] = row["Item Name"]
                            merged[f"{r} Price"] = row["New Price"]
                            merged[f"{r} Item SKU"] = row["Normalized Code"]
                            merged[f"{r} Product URL"] = row["Product URL"]
                        else:
                            merged[f"{r} Item Name"] = "N/A"
                            merged[f"{r} Price"] = pd.NA
                            merged[f"{r} Item SKU"] = row["Normalized Code"]
                            merged[f"{r} Product URL"] = "N/A"
                    merged["Confidence"] = 0.0
                    merged["Best Price"] = row["New Price"]
                    merged["Lowest Retailer"] = retailer
                    new_unmatched_rows.append(merged)
                    continue
                matches = process.extractOne(row["Item Name"], other_df["Item Name"].tolist(), scorer=fuzz.token_sort_ratio)
                if matches:
                    match_name, score, list_idx = matches
                    match_idx = other_df.index[list_idx]  # Get the DataFrame index
                    match_row = other_df.loc[match_idx]
                    if score >= HIGHLIGHT_CONFIDENCE_WEAK:
                        merged = pd.DataFrame({"Normalized Code": [row["Normalized Code"]]})
                        for r in RETAILER_FOLDERS:
                            if r == retailer:
                                merged[f"{r} Item Name"] = row["Item Name"]
                                merged[f"{r} Price"] = row["New Price"]
                                merged[f"{r} Item SKU"] = row["Normalized Code"]
                                merged[f"{r} Product URL"] = row["Product URL"]
                            elif r == match_row["Retailer"]:
                                merged[f"{r} Item Name"] = match_row["Item Name"]
                                merged[f"{r} Price"] = match_row["New Price"]
                                merged[f"{r} Item SKU"] = match_row["Normalized Code"]
                                merged[f"{r} Product URL"] = match_row["Product URL"]
                            else:
                                merged[f"{r} Item Name"] = "N/A"
                                merged[f"{r} Price"] = pd.NA
                                merged[f"{r} Item SKU"] = row["Normalized Code"]
                                merged[f"{r} Product URL"] = "N/A"
                        merged["Confidence"] = score
                        prices = {
                            r: merged[f"{r} Price"].values[0]
                            for r in RETAILER_FOLDERS if pd.notna(merged[f"{r} Price"].values[0])
                        }
                        best_price, best_retailer = (min(prices.values()), min(prices, key=prices.get)) if prices else (None, None)
                        merged["Best Price"] = best_price
                        merged["Lowest Retailer"] = best_retailer
                        weak_matched_rows.append(merged)
                        other_df = other_df.drop(index=match_idx)
                    else:
                        merged = pd.DataFrame({"Normalized Code": [row["Normalized Code"]]})
                        for r in RETAILER_FOLDERS:
                            if r == retailer:
                                merged[f"{r} Item Name"] = row["Item Name"]
                                merged[f"{r} Price"] = row["New Price"]
                                merged[f"{r} Item SKU"] = row["Normalized Code"]
                                merged[f"{r} Product URL"] = row["Product URL"]
                            else:
                                merged[f"{r} Item Name"] = "N/A"
                                merged[f"{r} Price"] = pd.NA
                                merged[f"{r} Item SKU"] = row["Normalized Code"]
                                merged[f"{r} Product URL"] = "N/A"
                        merged["Confidence"] = 0.0
                        merged["Best Price"] = row["New Price"]
                        merged["Lowest Retailer"] = retailer
                        new_unmatched_rows.append(merged)
                else:
                    merged = pd.DataFrame({"Normalized Code": [row["Normalized Code"]]})
                    for r in RETAILER_FOLDERS:
                        if r == retailer:
                            merged[f"{r} Item Name"] = row["Item Name"]
                            merged[f"{r} Price"] = row["New Price"]
                            merged[f"{r} Item SKU"] = row["Normalized Code"]
                            merged[f"{r} Product URL"] = row["Product URL"]
                        else:
                            merged[f"{r} Item Name"] = "N/A"
                            merged[f"{r} Price"] = pd.NA
                            merged[f"{r} Item SKU"] = row["Normalized Code"]
                            merged[f"{r} Product URL"] = "N/A"
                    merged["Confidence"] = 0.0
                    merged["Best Price"] = row["New Price"]
                    merged["Lowest Retailer"] = retailer
                    new_unmatched_rows.append(merged)
                other_df = other_df.reset_index(drop=True)  # Reset index after drop
    # Export files
    export_results(matched_rows, f"price-comparison-{category}-long-matched.xlsx", header_text=f"Price Comparison - {category} - Strong Matches")
    export_results(weak_matched_rows, f"price-comparison-{category}-long-weak-matched.xlsx", highlight_confidence={"threshold": HIGHLIGHT_CONFIDENCE_WEAK, "color": "FFFACD"})
    export_results(new_unmatched_rows, f"price-comparison-{category}-long-unmatched.xlsx", highlight_confidence={"threshold": HIGHLIGHT_CONFIDENCE_UNMATCHED, "color": "FF9999"})
    log(f"📊 {category}: Matched = {len(matched_rows)}, Weak Matched = {len(weak_matched_rows)}, Unmatched = {len(new_unmatched_rows)}")
    matched_categories.append(category)

# === Summary ===
skipped_list = {
    cat: list(srcs.keys())
    for cat, srcs in category_map.items()
    if len(srcs) < 2
}
log("\n✅ All comparisons completed.")
log(f"📁 Matched categories: {matched_categories}")
log(f"📁 Skipped categories: {list(skipped_list.keys())}")
if skipped_list:
    log("\n🚫 Skipped Categories (appear in only one retailer):")
    for cat, retailers in skipped_list.items():
        log(f"  - {cat} (from: {', '.join(retailers)})")
else:
    log("✅ No skipped categories due to missing retailer coverage.")